###################################################
#
# project: the scraper for https://hotcopper.com.au
# author: 
# date: 09/20/2019
#
###################################################

from pathlib import Path
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from requests.exceptions import ConnectionError
from requests.packages.urllib3.exceptions import MaxRetryError
from requests.packages.urllib3.exceptions import ProxyError as urllib3_ProxyError
#from xlwt import Workbook
from openpyxl import Workbook
import requests
import codecs
import time
import datetime
import sys
import os
import asyncio
import json
import random
import mysql.connector

proxy_base_url = "https://free-proxy-list.net/"
base_url = "https://hotcopper.com.au"
login_url = "https://hotcopper.com.au/login/"
asx_stock_url = "https://hotcopper.com.au/discussions/asx---by-stock/?post_view=0"
page_url = "https://hotcopper.com.au/discussions/asx---by-stock/page-{}"
username = "XXXX@outlook.com"
userpassword = "123456"

g_proxylist = []

def get_proxylist():
  global g_proxylist
  try:
    session = requests.Session()
    result = session.get(proxy_base_url)
    soup = BeautifulSoup(result.content, 'html.parser')
    table_arr = soup.select('table#proxylisttable')
    if len(table_arr) > 0:
      table_tbody_arr = table_arr[0].select('tbody')
      if len(table_tbody_arr) > 0:
        table_tbody = table_tbody_arr[0]
        tr_arr = table_tbody.select('tr')
        print("-:proxy count", len(tr_arr))
        for tr_body in tr_arr:
          td_arr = tr_body.select('td')
          if len(td_arr) == 8:
            _proxy_ip = td_arr[0].text
            _proxy_port = td_arr[1].text
            _proxy_flag = "N"
            if td_arr[6].text == "yes":
              _proxy_flag = "Y"

            proxy_node = {"url": _proxy_ip + ":" + _proxy_port, "https": _proxy_flag}
            print(proxy_node)
            g_proxylist.append(proxy_node)

  except ConnectionError as ce:
    if (isinstance(ce.args[0], MaxRetryError) and isinstance(ce.args[0].reason, urllib3_ProxyError)):
      print("unable to connect to https://free-proxy-list.net/.")

pathname = os.path.dirname(sys.argv[0])

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  passwd="",
  database="hotcopper"
)
mycursor = mydb.cursor()

print("please select following options.(1 or 2)")
print("-1: export into excel file.")
print("-2: get the post data.")
input_param = input("option:")

#export into excel file
if input_param == "1":
  q_offset = 0
  q_limit = 0
  excel_file = input("save to:")
  start_index = input("start index:")
  export_count = input("count to export:")
  if excel_file == "":
    excel_file = "default.xlsx"
  if start_index != "":
    q_offset = int(start_index)
  if export_count != "":
    q_limit = int(export_count)

  g_ExcelPath = pathname + "\\" + excel_file
  print("output file:", g_ExcelPath)

  g_excel_number = 1
  wb = Workbook()
  excel_sheet = wb.active
  excel_sheet.title = "Reviews"
  excel_sheet.cell(row=1, column=1).value = "Name"
  excel_sheet.cell(row=1, column=2).value = "Number_Posts"
  excel_sheet.cell(row=1, column=3).value = "Lightbulb"
  excel_sheet.cell(row=1, column=4).value = "Date_Time"
  excel_sheet.cell(row=1, column=5).value = "Post #"
  excel_sheet.cell(row=1, column=6).value = "Commentary"
  excel_sheet.cell(row=1, column=7).value = "Images_Commentary"
  excel_sheet.cell(row=1, column=8).value = "Ticker"
  excel_sheet.cell(row=1, column=9).value = "Price at Posting"
  excel_sheet.cell(row=1, column=10).value = "Sentiment"
  excel_sheet.cell(row=1, column=11).value = "Disclosure"

  wb.save(g_ExcelPath)

  q_offset_str = ""
  q_limit_str = ""
  if q_offset >= 0:
    q_offset_str = " OFFSET " + str(q_offset) + " "
  if q_limit > 0:
    q_limit_str = " LIMIT " + str(q_limit) + " "

  sql = "SELECT name, post_count, ga_count, post_date, post_no, commentary, images_commentary, ticker, post_price, post_sentiment, disclosure"
  sql = sql + " FROM postlist " + q_limit_str + q_offset_str
  mycursor.execute(sql)
  myresult = mycursor.fetchall()
  for node in myresult:
    _name = node[0]
    _post_count = node[1]
    _ga_count = node[2]
    _post_date = node[3]
    _post_no = node[4]
    _commentary = node[5]
    _image_commentary = node[6]
    _ticker = node[7]
    _post_price = node[8]
    _post_sentiment = node[9]
    _disclosure = node[10]

    g_excel_number = g_excel_number + 1
    try:
      excel_sheet.cell(row=g_excel_number, column=1).value = _name
      excel_sheet.cell(row=g_excel_number, column=2).value = _post_count
      excel_sheet.cell(row=g_excel_number, column=3).value = _ga_count
      excel_sheet.cell(row=g_excel_number, column=4).value = _post_date
      excel_sheet.cell(row=g_excel_number, column=5).value = _post_no
      excel_sheet.cell(row=g_excel_number, column=6).value = _commentary
      excel_sheet.cell(row=g_excel_number, column=7).value = _image_commentary
      excel_sheet.cell(row=g_excel_number, column=8).value = _ticker
      excel_sheet.cell(row=g_excel_number, column=9).value = _post_price
      excel_sheet.cell(row=g_excel_number, column=10).value = _post_sentiment
      excel_sheet.cell(row=g_excel_number, column=11).value = _disclosure
    except:
      print("--:Error", _post_no, _ticker, _name)

    wb.save(g_ExcelPath)

  wb.close()
  print("file was saved to ", g_ExcelPath)

#get the post data.
if input_param == "2":
  get_proxylist()

  options = Options()
  options.add_argument('--log-level=3')
  options.add_argument("--disable-extensions")
  options.add_argument("--incognito")
  selected_proxy_index = -1

  #set proxy in chromedriver
  if len(g_proxylist) > 0:
    selected_proxy_index = random.randint(0, len(g_proxylist) - 1)
  if selected_proxy_index > 0:
    options.add_argument("'--proxy-server={}".format(g_proxylist[selected_proxy_index]['url']))

  driver = webdriver.Chrome('chromedriver', options=options)
  driver.get(login_url)
  driver.implicitly_wait(3)
  time.sleep(3)

  #click username and input
  driver.find_element_by_xpath("//input[@name='login']").click()
  time.sleep(1)
  driver.find_element_by_xpath("//input[@name='login']").send_keys(username)
  time.sleep(1)

  #click username and input
  driver.find_element_by_xpath("//input[@name='password']").click()
  time.sleep(1)
  driver.find_element_by_xpath("//input[@name='password']").send_keys(userpassword)
  time.sleep(1)

  #click siginin button
  driver.find_element_by_xpath("//label[@for='tos']").click()
  time.sleep(3)

  #click siginin button
  driver.find_element_by_xpath("//button[@id='btn-login']").click()
  time.sleep(3)

  driver.get(asx_stock_url)
  time.sleep(3)

  page_num = 1
  open_link = asx_stock_url
  ticktime = 3
  while True:
    if page_num > 1:
      driver.get(open_link)
      ticktime = random.randint(3, 6)
      time.sleep(ticktime)

    s = BeautifulSoup(driver.page_source, 'html.parser')
    post_arr = s.select('tr.thread-tr')
    kk = 0
    for post_node in post_arr:
      kk = kk + 1
      bb = divmod(kk, 2)
      if bb[1] == 0:
        continue
      sublink = ""
      postlink = ""
      subject = ""
      replies = ""
      tag = ""
      company_name = ""

      #get subject, code of post
      postlink_arr = post_node.select('td.subject-td > strong > a')
      if len(postlink_arr) > 0:
        postlink = postlink_arr[0]["href"]
        subject = postlink_arr[0].text
        sublink = base_url + postlink
        sublink = sublink.replace("/unread", "/")
        code_arr = sublink.split(".")
        code_len = len(code_arr)
        code = code_arr[code_len - 1].replace("/", "")

      #get replies
      replies_arr = post_node.select('td.replies-td')
      if len(replies_arr) > 0:
        replies = replies_arr[0].text

      #get company tag
      stockpill_arr = post_node.select('td.stock-pill-td')
      if len(stockpill_arr) > 0:
        tag = stockpill_arr[0].text

      #get company title
      stockpilltag_arr = post_node.select('td.stock-pill-td > span.stock-pill > a')
      if len(stockpilltag_arr) > 0:
        company_title = stockpilltag_arr[0]["title"]
        company_title_arr = company_title.split("-")
        company_name = company_title_arr[0].strip()

      #register company
      sql = "SELECT * FROM company WHERE asx_code='" + tag + "'"
      mycursor.execute(sql)
      myresult = mycursor.fetchone()
      if mycursor.rowcount <= 0:
        sql = "INSERT INTO company(company_name, asx_code)VALUES(%s, %s)"
        val = (company_name, tag)
        mycursor.execute(sql, val)
      else:
        print("company already is exist.", tag)

      print(sublink, subject, code, replies, tag, company_name)

      #check to already exist post
      sql = "SELECT id, replies, link FROM post_company WHERE post_no='"+ code + "'"
      mycursor.execute(sql)
      myresult = mycursor.fetchone()
      bexist = False
      if mycursor.rowcount > 0:
        bexist = True
        last_id = myresult[0]
        old_replies = myresult[1]
        if old_replies == replies:
          print("same post already is exists.", code)
          continue
        else:
          print("the post was changed.", code)
          first_link = myresult[2]
          sql  = "UPDATE post_company SET replies='" + replies + "' WHERE id=" + str(last_id)
          mycursor.execute(sql)

      if bexist == False:
        sql = "INSERT INTO post_company(tag, post_no, subject, replies, link) VALUES (%s, %s, %s, %s, %s)"
        val = (tag, code, subject, replies, sublink)
        mycursor.execute(sql, val)
        last_id = mycursor.lastrowid
        mydb.commit()

        first_link = sublink
        subpage_no = 1

        while True:
          driver.get(first_link)
          ticktime = random.randint(3, 6)
          time.sleep(ticktime)
          current_url = driver.current_url
          if first_link != current_url:
            print("this is last pages.")
            break

          ss = BeautifulSoup(driver.page_source, 'html.parser')
          #get post list
          message_arr = ss.select('div.message-columns')
          for message_node in message_arr:
            msg_username = ""
            msg_postno = ""
            msg_postcount = ""
            msg_gacount = ""
            msg_postdate = ""
            msg_priceposition = ""
            msg_sentiment = ""
            msg_disclosure = ""
            msg_content = ""

            #get post number
            msg_postno_arr = message_node.select('div.post-link > a')
            if len(msg_postno_arr) > 0:
              msg_postno = msg_postno_arr[0].text

            #check to exist post
            sql = "SELECT * FROM postlist WHERE post_no='" + msg_postno + "'"
            mycursor.execute(sql)
            myresult = mycursor.fetchone()
            if mycursor.rowcount > 0:
              print("exist post", msg_postno)
              continue

            #get username
            msg_username_arr = message_node.select('div.user-username > a')
            if len(msg_username_arr) > 0:
              msg_username = msg_username_arr[0].text

            #get post count
            msg_postcount_arr = message_node.select('div.user-post-num')
            if len(msg_postcount_arr) > 0:
              msg_postcount = msg_postcount_arr[0].text
              msg_postcount = msg_postcount.replace("\"", "", 10)
              msg_postcount = msg_postcount.replace("\n", "", 10)
              msg_postcount = msg_postcount.replace("Posts.", "").strip()

            #get ga count
            msg_gacount_arr = message_node.select('div.user-ga-count')
            if len(msg_gacount_arr) > 0:
              msg_gacount = msg_gacount_arr[0].text
              msg_gacount = msg_gacount.replace("lightbulb Created with Sketch.", "").strip()

            #get post date
            msg_postdate_arr = message_node.select('div.post-metadata-date')
            if len(msg_postdate_arr) > 0:
              msg_postdate = msg_postdate_arr[0].text
              msg_postdate = msg_postdate.replace("\n", "", 10).strip()

            #get post content
            msg_content_arr = message_node.select('div.message-content > article')
            if len(msg_content_arr) > 0:
              msg_content = msg_content_arr[0].text

            #get price, sentiment and disclosure
            msg_metadata_arr = message_node.select('div.message-content > div.message-user-metadata > div > span')
            if len(msg_metadata_arr) >= 4:
              msg_priceposition = msg_metadata_arr[1].text
              msg_priceposition = msg_priceposition.replace("Price at posting:", "")
              msg_priceposition = msg_priceposition.replace("\"", "", 10).strip()

              msg_sentiment = msg_metadata_arr[2].text
              msg_sentiment = msg_sentiment.replace("Sentiment:", "").strip()

              msg_disclosure = msg_metadata_arr[3].text
              msg_disclosure = msg_disclosure.replace("Disclosure:", "").strip()

            print("-:username=", msg_username, "postno=", msg_postno, "postcount=", msg_postcount,
                  "gacount=", msg_gacount, "postdate=", msg_postdate, "priceposition=", msg_priceposition,
                  "sentiment=", msg_sentiment, "disclosure=", msg_disclosure)
            #print("msg_content=", msg_content)

            #insert data to database
            sql = "INSERT INTO postlist(company_id, name, post_count, ga_count, post_date, post_no, commentary, ticker, post_price, post_sentiment, disclosure)"
            sql = sql + "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            val = (last_id, msg_username, msg_postcount, msg_gacount, msg_postdate, msg_postno, msg_content, tag, msg_priceposition, msg_sentiment, msg_disclosure)
            mycursor.execute(sql, val)
            mydb.commit()

          subpage_no = subpage_no + 1
          first_link = sublink + "page-" + str(subpage_no)

    page_num = page_num + 1
    open_link = page_url.format(page_num)


  time.sleep(5)
  driver.close()

print("THE END")
exit(0)
